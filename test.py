import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment

# Danh sách task: [Tên, thời gian bắt đầu, thời lượng]
tasks = [
    ['A', 0, 5],
    ['B', 5, 5],
    ['C1', 10, 2],
    ['C2', 10, 2],
    ['C3', 10, 2],
    ['C4', 10, 14],
    ['D', 24, 14],
    ['E2', 12, 3],
    ['E3', 15, 14],
    ['E4', 29, 7],
    ['E1', 38, 1],
    ['G1', 38, 2],
    ['G2', 40, 2],
    ['G3', 42, 3],
    ['H1', 39, 10],
    ['H2', 49, 15],
    ['H3', 64, 42],
    ['I2', 49, 15],
    ['J2', 64, 11],
    ['K2', 75, 13],
    ['L', 88, 8],
    ['M', 96, 1],
    ['K', 106, 5],
]

# Tạo bảng màu ngẫu nhiên cho từng task
def random_color():
    return ''.join(random.choices('ABCDEF0123456789', k=6))

color_map = {
    task[0]: PatternFill(
        start_color=random_color(),
        end_color=random_color(),
        fill_type="solid"
    )
    for task in tasks
}

# Tạo workbook và worksheet
wb = Workbook()
ws = wb.active
ws.title = "Gantt Chart"

# Header: Thời gian
ws.cell(row=1, column=1, value="Task")
max_time = max(start + duration for _, start, duration in tasks)
for t in range(max_time + 1):
    cell = ws.cell(row=1, column=t + 2, value=t)
    col_letter = cell.column_letter
    ws.column_dimensions[col_letter].width = 3  # Thu nhỏ cell theo chiều ngang

# Ghi dữ liệu task
for i, (task, start, duration) in enumerate(tasks):
    row = i + 2
    ws.cell(row=row, column=1, value=task)  # Tên task ở cột đầu
    for t in range(start, start + duration):
        cell = ws.cell(row=row, column=t + 2, value=task)  # In tên task vào ô
        cell.fill = color_map[task]
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Giảm chiều cao các dòng
for i in range(2, len(tasks) + 2):
    ws.row_dimensions[i].height = 15  # tùy chỉnh chiều cao dòng

# Lưu file
wb.save("gantt_chart_colored_named.xlsx")
